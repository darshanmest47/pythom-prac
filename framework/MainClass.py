import inspect
import logging

import openpyxl
import pytest

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@pytest.mark.usefixtures("setup")
class MainClass:


    def waituntillpresent(self,driver):
        waits=WebDriverWait(driver,20)
        return waits

    def signSelect(self,locator,content):
        sel=Select(locator)
        sel.select_by_visible_text(content)

    def loggerFunc(self):
        name=inspect.stack()[1][3]
        logger=logging.getLogger(name)
        filehandler=logging.FileHandler('files.log')
        formatter=logging.Formatter('%(asctime)s : %(name)s : %(levelname)s : %(message)s')
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)

        logger.setLevel('DEBUG')

        return logger

    def getExceldata(self,testdataname):
        path="E:\\pythom-prac\\framework\\Excel\\Testdata.xlsx"
        wb=openpyxl.load_workbook(path)
        sheet=wb.active
        mxrow=sheet.max_row
        mxcol=sheet.max_column
        dicts={}
        listsdata=[]

        for ele in range(len(testdataname)):


            for i in range(1,mxrow+1):
                 if sheet.cell(row=i,column=1).value==testdataname[ele]:
                     print(testdataname[ele])

                     for j in range(2, mxcol + 1):
                         print(sheet.cell(row=i, column=j).value)
                         dicts[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

                     print(dicts)
                     listsdata.append(dicts)
                     dicts = {}






        print(listsdata)
        return listsdata



















